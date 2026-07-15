"""通用 Excel 导出工具

封装 openpyxl + StreamingResponse,供 batch/detection/alert 导出复用。
用法:export_to_excel(columns, rows, filename) -> StreamingResponse
  - columns: [(字段名, 表头), ...] 表头即列标题
  - rows: [dict, ...] 每行数据(按字段名取值)
"""
import io
from datetime import datetime
from typing import Any, List, Tuple

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def export_to_excel(
    columns: List[Tuple[str, str]],
    rows: List[dict],
    filename: str,
) -> StreamingResponse:
    """生成 Excel 文件并返回 StreamingResponse 供下载。

    columns: [(key, header), ...] —— key 用于从 row 取值,header 是表头
    rows: [dict, ...]
    filename: 不含扩展名的文件名,自动加日期后缀和 .xlsx
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "数据"

    # 表头样式
    header_fill = PatternFill(start_color="7CFF67", end_color="7CFF67", fill_type="solid")
    header_font = Font(bold=True, color="0F3D22")
    center = Alignment(horizontal="center", vertical="center")

    # 写表头
    headers = [h for _, h in columns]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # 写数据行
    for row in rows:
        ws.append([_cell_value(row.get(k)) for k, _ in columns])

    # 自适应列宽(粗略)
    for idx, (_, header) in enumerate(columns, start=1):
        max_len = len(str(header))
        for row in rows:
            val = row.get(columns[idx - 1][0])
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = min(max_len + 4, 40)

    # 写到内存
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    date_str = datetime.now().strftime("%Y%m%d")
    fname = f"{filename}_{date_str}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


def _cell_value(val: Any) -> Any:
    """处理 None / datetime 等类型,openpyxl 不支持的部分转字符串"""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M:%S")
    return val
