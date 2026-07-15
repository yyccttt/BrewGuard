import csv
import io
from datetime import datetime

from fastapi import APIRouter, Query, UploadFile, File, Form
from tortoise.expressions import Q

from app.controllers.detection import detection_controller
from app.core.websocket import manager
from app.models.alerts import Alert
from app.models.batch import Batch
from app.models.detection import Detection
from app.schemas import Success, SuccessExtra
from app.schemas.detections import *
from app.utils.export import export_to_excel

router = APIRouter()

# 默认阈值(批次未配置时使用)
DEFAULT_THRESHOLDS = {
    "temp_min": 18, "temp_max": 28,
    "ph_min": 3.5, "ph_max": 4.5,
    "abv_min": 0, "abv_max": 15,
}


async def check_and_alert(batch_id: int, detection_id: int, temperature, ph, abv):
    """检查检测值是否超阈值,超了则创建告警,并通过 WebSocket 实时推送"""
    batch = await Batch.filter(id=batch_id).first()
    if not batch:
        return

    def threshold(field):
        val = getattr(batch, field, None)
        return val if val is not None else DEFAULT_THRESHOLDS[field]

    checks = [
        ("temperature", temperature, threshold("temp_min"), threshold("temp_max")),
        ("ph", ph, threshold("ph_min"), threshold("ph_max")),
        ("abv", abv, threshold("abv_min"), threshold("abv_max")),
    ]

    new_alerts = []
    for metric, value, lo, hi in checks:
        if value is None:
            continue
        if value > hi:
            alert = await Alert.create(
                batch_id=batch_id, detection_id=detection_id, metric=metric,
                value=value, threshold=hi, direction="high", status="open"
            )
            new_alerts.append((alert, metric, value, hi, "high"))
        elif value < lo:
            alert = await Alert.create(
                batch_id=batch_id, detection_id=detection_id, metric=metric,
                value=value, threshold=lo, direction="low", status="open"
            )
            new_alerts.append((alert, metric, value, lo, "low"))

    # 实时推送告警给所有在线客户端
    for alert, metric, value, threshold_val, direction in new_alerts:
        await manager.broadcast_json({
            "type": "alert",
            "payload": {
                "id": alert.id, "batch_id": batch_id, "batch_no": batch.batch_no,
                "metric": metric, "value": value, "threshold": threshold_val,
                "direction": direction, "status": "open",
                "created_at": str(alert.created_at) if hasattr(alert, "created_at") else None,
            },
        })


@router.get("/list", summary="查看检测记录列表")
async def list_detection(
    page: int = Query(1, description="页码"),
    page_size: int = Query(20, description="每页数量"),
    batch_id: int = Query(..., description="批次ID"),
):
    q = Q(is_deleted=False) & Q(batch_id=batch_id)
    total, objs = await detection_controller.list(page=page, page_size=page_size, search=q)
    data = [await obj.to_dict() for obj in objs]
    return SuccessExtra(data=data, total=total, page=page, page_size=page_size)


@router.get("/get", summary="查看检测记录")
async def get_detection(
    id: int = Query(..., description="检测记录ID"),
):
    obj = await detection_controller.get(id=id)
    data = await obj.to_dict()
    return Success(data=data)


@router.post("/create", summary="创建检测记录")
async def create_detection(detection_in: DetectionCreate):
    obj = await detection_controller.create(obj_in=detection_in)
    # 推送实时检测数值给在线客户端(仪表盘实时跳动)
    batch = await Batch.filter(id=detection_in.batch_id).first()
    await manager.broadcast_json({
        "type": "detection",
        "payload": {
            "id": obj.id, "batch_id": detection_in.batch_id,
            "batch_no": batch.batch_no if batch else None,
            "temperature": detection_in.temperature, "ph": detection_in.ph,
            "abv": detection_in.abv, "remark": detection_in.remark,
        },
    })
    # 自动检查阈值并触发告警(内部会推送告警消息)
    await check_and_alert(detection_in.batch_id, obj.id, detection_in.temperature, detection_in.ph, detection_in.abv)
    return Success(msg="Created Successfully")


@router.post("/update", summary="更新检测记录")
async def update_detection(detection_in: DetectionUpdate):
    await detection_controller.update(id=detection_in.id, obj_in=detection_in)
    return Success(msg="Updated Successfully")


@router.delete("/delete", summary="删除检测记录")
async def delete_detection(
    id: int = Query(..., description="检测记录ID"),
):
    obj = await detection_controller.get(id=id)
    obj.is_deleted = True
    await obj.save()
    return Success(msg="Deleted Successfully")


@router.get("/export", summary="导出某批次检测记录 Excel (#40)")
async def export_detection(batch_id: int = Query(..., description="批次ID")):
    objs = await Detection.filter(is_deleted=False, batch_id=batch_id).order_by("created_at")
    rows = [await obj.to_dict() for obj in objs]
    columns = [
        ("temperature", "温度(°C)"), ("ph", "pH"), ("abv", "酒精度(%)"),
        ("remark", "备注"), ("created_at", "记录时间"),
    ]
    return export_to_excel(columns, rows, f"detection_batch_{batch_id}")


async def _parse_rows_from_file(file: UploadFile):
    """解析上传的 csv/xlsx,返回 [{temperature, ph, abv, remark}, ...]"""
    content = await file.read()
    filename = (file.filename or "").lower()
    rows = []
    if filename.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        for r in reader:
            rows.append(r)
    elif filename.endswith((".xlsx", ".xls")):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(max_row=1))]
        for r in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in r):
                continue
            rows.append({headers[i]: v for i, v in enumerate(r) if i < len(headers)})
    else:
        raise ValueError("仅支持 .csv / .xlsx 文件")
    return rows


def _to_float(val):
    """宽松转 float,空值/非法值返回 None"""
    if val is None or val == "":
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


@router.post("/import", summary="批量导入检测记录 (#42)")
async def import_detection(
    batch_id: int = Form(..., description="批次ID"),
    file: UploadFile = File(..., description="csv/xlsx 文件"),
):
    """批量导入检测记录,逐条触发告警检查。返回成功/失败计数 + 错误明细"""
    try:
        rows = await _parse_rows_from_file(file)
    except ValueError as e:
        return Success(msg=f"解析失败: {e}", data={"success_count": 0, "fail_count": 0, "errors": [str(e)]})

    success_count = 0
    fail_count = 0
    errors = []
    for idx, r in enumerate(rows, start=2):  # start=2 对应表格第2行(跳过表头)
        try:
            temperature = _to_float(r.get("temperature"))
            ph = _to_float(r.get("ph"))
            abv = _to_float(r.get("abv"))
            remark = str(r.get("remark") or "")
            obj = await Detection.create(
                batch_id=batch_id, temperature=temperature, ph=ph, abv=abv, remark=remark,
            )
            # 复用告警检查 + 推送
            await check_and_alert(batch_id, obj.id, temperature, ph, abv)
            success_count += 1
        except Exception as e:
            fail_count += 1
            errors.append(f"第{idx}行: {e}")

    return Success(msg=f"导入完成: 成功{success_count}条, 失败{fail_count}条", data={
        "success_count": success_count, "fail_count": fail_count, "errors": errors[:20],
    })


@router.get("/template", summary="下载检测记录导入模板 (#42)")
async def download_template():
    """下载标准导入模板 .xlsx"""
    columns = [
        ("temperature", "温度(°C)"), ("ph", "pH"), ("abv", "酒精度(%)"), ("remark", "备注"),
    ]
    return export_to_excel(columns, [], "detection_import_template")
